import os
import json
import urllib.request
import urllib.error
import openpyxl

def sync():
    excel_path = "c:\\toyzgurunewweb\\toyzguru_products_catalog.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return

    print(f"Reading Excel catalog from: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    
    headers = [sheet.cell(1, c).value for c in range(1, 14)]
    print(f"Headers found: {headers}")

    products = []
    for r in range(2, sheet.max_row + 1):
        row_id = sheet.cell(r, 1).value
        if not row_id:
            continue
            
        row_vals = {headers[c-1]: sheet.cell(r, c).value for c in range(1, 14)}
        
        # Format options
        opts_val = row_vals.get('options')
        if not opts_val:
            options_arr = []
        elif isinstance(opts_val, str):
            options_arr = [o.strip() for o in opts_val.split(',') if o.strip()]
        else:
            options_arr = [str(opts_val)]
            
        # Format specs
        specs_val = row_vals.get('specs')
        if not specs_val:
            specs_obj = {}
        elif isinstance(specs_val, str):
            try:
                specs_obj = json.loads(specs_val)
            except Exception:
                specs_obj = {"Specs": specs_val}
        else:
            specs_obj = specs_val
            
        # Build product representation matching Supabase schema
        prod = {
            "id": str(row_vals['id']),
            "title": str(row_vals['title']),
            "category": str(row_vals['category']),
            "price": float(row_vals['price']),
            "original_price": float(row_vals['original_price']) if row_vals.get('original_price') is not None else None,
            "image": str(row_vals['image']),
            "rating": float(row_vals['rating']) if row_vals.get('rating') is not None else 5.0,
            "reviews_count": int(row_vals['reviews_count']) if row_vals.get('reviews_count') is not None else 1,
            "badge": str(row_vals['badge']) if row_vals.get('badge') is not None else None,
            "description": str(row_vals['description']) if row_vals.get('description') is not None else "",
            "options": options_arr,
            "specs": specs_obj,
            "stock": int(row_vals['stock']) if row_vals.get('stock') is not None else 10
        }
        products.append(prod)

    print(f"Parsed {len(products)} products from Excel.")

    # Supabase REST endpoint
    url = "https://lunwguzzguemtotsshjm.supabase.co/rest/v1/products?on_conflict=id"
    headers = {
        "apikey": "sb_publishable_E9_XvSijoHdGk83Iv09wNg_DjO1B9EX",
        "Authorization": "Bearer sb_publishable_E9_XvSijoHdGk83Iv09wNg_DjO1B9EX",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates"
    }

    # Perform Upsert
    data_bytes = json.dumps(products).encode('utf-8')
    req = urllib.request.Request(url, data=data_bytes, headers=headers, method="POST")
    
    try:
        with urllib.request.urlopen(req) as response:
            status = response.getcode()
            print(f"Supabase returned status code: {status}")
            print("Successfully synced all products to Supabase!")
    except urllib.error.HTTPError as e:
        print(f"HTTP Error: {e.code} - {e.read().decode('utf-8')}")
    except Exception as e:
        print(f"Error connecting to Supabase: {e}")

if __name__ == "__main__":
    sync()
